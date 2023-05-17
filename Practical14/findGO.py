from openpyxl import Workbook
import xml.dom.minidom
#First, parse the XML file into a DOM document
DOMTree = xml.dom.minidom.parse('go_obo.xml')
# Get all the <term> elements
all_elements = DOMTree.getElementsByTagName('term')

#!!Excel 1.workbook 2.active sheets 3.write in sheets
workbook = Workbook()
sheet = workbook.active
sheet['A1'] = 'id'
sheet['B1'] = 'name'
sheet['C1'] = 'definition'
sheet['D1'] = 'childnodes'
# row_index = 2 for writing data to the sheet (!!! the first row for headers)
row_index = 2

# Most difficult!!! count the number of child nodes
def count_childnodes(termid):
    nodes = 0
    for subterm in all_elements:
        is_a_list = subterm.getElementsByTagName('is_a') 
        for is_a_instance in is_a_list:
            if termid == is_a_instance.firstChild.data:
                nodes += 1 + count_childnodes(subterm.getElementsByTagName('id')[0].firstChild.data)
    return nodes

for term_element in all_elements:
    # definition
    definition = term_element.getElementsByTagName('defstr')[0].firstChild.data
    
    # If the text in <defstr> contains 'autophagosome'+ IMPORTANT: variable.firstChild.data
    if 'autophagosome' in definition:
        go_id = term_element.getElementsByTagName('id')[0].firstChild.data
        term_name = term_element.getElementsByTagName('name')[0].firstChild.data

        # Write intoto the Excel sheet
        sheet.cell(row=row_index, column=1).value = go_id
        sheet.cell(row=row_index, column=2).value = term_name
        sheet.cell(row=row_index, column=3).value = definition
        sheet.cell(row=row_index, column=4).value = count_childnodes(go_id)
        row_index += 1 #incremnet each time


# Save the Excel workbook
workbook.save('autophagosome.xlsx')